import streamlit as st
import pandas as pd
import re
import pdfplumber
import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from dateutil.parser import parse as du_parse
from datetime import datetime

# --- CONFIG & STYLES ---
st.set_page_config(page_title="Hotel Reconcile Pro (v1.15 Web)", page_icon="🏨", layout="wide")

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

MONTH_ABBR = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
MONTH_MAP  = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

# --- ORIGINAL HELPERS (From v1.15) ---

def norm_key(v):
    s = "" if v is None else str(v)
    s = s.strip().upper()
    s = re.sub(r"[ \-]+", "", s)
    return s

def infer_hotel(text: str) -> str:
    t = (text or "").lower()
    norm = re.sub(r"[^a-z0-9]+", " ", t).strip()
    pad  = f" {norm} "
    if "katathani" in norm: return "KT"
    if "the shore" in norm: return "TS"
    if "the waters" in norm or "waters" in norm: return "WAT"
    if "the little shore" in norm or "little shore" in norm: return "TLKL"
    if "the sands" in norm or "sands" in norm: return "SAN"
    if "the leaf on the sands" in norm or "leaf on the sands" in norm: return "LFS"
    if "the leaf oceanside" in norm or "leaf oceanside" in norm: return "LFO"
    if " kt "   in pad: return "KT"
    if " ts "   in pad: return "TS"
    if " wat "  in pad: return "WAT"
    if " tlkl " in pad: return "TLKL"
    if " san "  in pad: return "SAN"
    if " lfs "  in pad: return "LFS"
    if " lfo "  in pad: return "LFO"
    return "UNK"

def infer_ota(text):
    t = (text or "").lower()
    if "booking" in t: return "Booking.com"
    if "expedia" in t or "hotels.com" in t: return "Expedia"
    if "agoda" in t: return "Agoda"
    if "traveloka" in t: return "Traveloka"
    if "trip.com" in t or "ctrip" in t: return "Trip.com"
    return "OTA"

def canon_ota(name):
    n = (name or "").lower()
    n = re.sub(r"[^a-z0-9]", "", n)
    if any(x in n for x in ["bookingcom","booking"]): return "Booking.com"
    if any(x in n for x in ["expedia","hotelscom","hotels"]): return "Expedia"
    return n

def monyy(y:int, m:int): 
    return f"{MONTH_ABBR.get(m,'Mon')}'{str(y)[-2:]}"

def extract_period_from_name(name):
    if not name: return None
    n = name.lower()
    m = re.search(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s*'(\d{2})\b", n)
    if m: return monyy(int("20"+m.group(2)), MONTH_MAP[m.group(1)])
    m = re.search(r"\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?)\D{0,3}(20\d{2})\b", n)
    if m: return monyy(int(m.group(2)), MONTH_MAP[m.group(1)[:3]])
    m = re.search(r"\b(20\d{2})[ _./-](0[1-9]|1[0-2])\b", n)
    if m: return monyy(int(m.group(1)), int(m.group(2)))
    m = re.search(r"\b(0[1-9]|1[0-2])[ _./-](20\d{2})\b", n)
    if m: return monyy(int(m.group(2)), int(m.group(1)))
    return None

def choose_period(hot_name, com_name, hot_earliest=None, com_earliest=None):
    p = extract_period_from_name(com_name) or extract_period_from_name(hot_name)
    if not p:
        def fmt(dt): return dt.strftime("%b'%y") if isinstance(dt, datetime) else None
        p = fmt(com_earliest) or fmt(hot_earliest) or "Period"
    return p

def trim_sheet(n):
    n = re.sub(r'[\\/*?:\\[\\]]', "_", n)
    return n[:31]

def find_header_col(ws, header_names, header_row=2):
    names = set(h.lower() for h in header_names)
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=header_row, column=c).value
        s = "" if v is None else str(v).strip().lower()
        if s in names: return c
    return None

def format_dates(ws, cols, start_row, end_row):
    for col in cols:
        if col is None: continue
        for r in range(start_row, end_row+1):
            ws.cell(row=r, column=col).number_format = "yyyy-mm-dd"

# --- CORE LOGIC: Hoteliers Processing ---

def collect_hoteliers(ws):
    header_row = 2
    res_col = find_header_col(ws, ["Reservation number","Reservation id","Confirmation number"], header_row) or 2
    arr_col = find_header_col(ws, ["Arrival","Check-in","Check in"], header_row)
    dep_col = find_header_col(ws, ["Departure","Check-out","Check out"], header_row)
    ota_col = find_header_col(ws, [
        "Channel","OTA","Booking Channel","Source","Distributor","Partner","Agency","Agent",
        "Merchant","Reservation Source","Booking Site","Website","Booking Source"
    ], header_row)

    last_nonempty = ws.max_row
    while last_nonempty > 1:
        if any(ws.cell(row=last_nonempty, column=c).value not in (None, "") for c in range(1, ws.max_column+1)):
            break
        last_nonempty -= 1
    data_start = 3
    data_end = max(2, last_nonempty - 1)

    keys_by_period, rows_by_period = {}, {}
    earliest_dep = None

    for r in range(data_start, data_end+1):
        d = ws.cell(row=r, column=dep_col).value if dep_col else None
        dt = d if isinstance(d, datetime) else None
        if not dt and d not in (None, ""):
            try: dt = du_parse(str(d), fuzzy=True)
            except Exception: dt = None
        if dt is None and arr_col:
            a = ws.cell(row=r, column=arr_col).value
            if isinstance(a, datetime): dt = a
            elif a not in (None, ""):
                try: dt = du_parse(str(a), fuzzy=True)
                except Exception: dt = None
        
        if dt:
            period = monyy(dt.year, dt.month)
            if earliest_dep is None or dt < earliest_dep: earliest_dep = dt
        else:
            period = "Unknown"
        
        k = norm_key(ws.cell(row=r, column=res_col).value)
        if k:
            keys_by_period.setdefault(period, set()).add(k)
            rows_by_period.setdefault(period, []).append(r)

    format_dates(ws, [arr_col, dep_col], data_start, data_end)
    return {"res_col": res_col, "rows_by_period": rows_by_period, "keys_by_period": keys_by_period,
            "earliest_dep": earliest_dep, "arr_col": arr_col, "dep_col": dep_col, "ota_col": ota_col}

# --- CORE LOGIC: OTA Excel Processing ---

def detect_commission_settings(ws):
    # Auto-detect headers logic from v1.15
    keywords = ["reservation", "booking id"]
    best_row, best_col = 1, 2
    
    # Simple scan first 10 rows
    for r in range(1, min(11, ws.max_row+1)):
        for c in range(1, ws.max_column+1):
            v = str(ws.cell(row=r, column=c).value or "").lower()
            if any(k in v for k in keywords):
                best_row = r
                best_col = c
                break
        if best_row != 1: break
        
    return {"header_row": best_row, "reservation_col_idx": best_col, "data_start_row": best_row + 1}

def collect_commission_excel(ws, settings):
    col_idx = settings["reservation_col_idx"]
    data_start = settings["data_start_row"]
    keys = set()
    earliest = None
    
    # Try to find date column for earliest date detection
    hdr = settings["header_row"]
    date_col = None
    for c in range(1, ws.max_column+1):
        v = str(ws.cell(row=hdr, column=c).value or "").lower()
        if "arrival" in v or "check-in" in v or "departure" in v:
            date_col = c
            break
            
    for r in range(data_start, ws.max_row+1):
        k = norm_key(ws.cell(row=r, column=col_idx).value)
        if k: keys.add(k)
        if date_col:
            val = ws.cell(row=r, column=date_col).value
            dt = val if isinstance(val, datetime) else None
            if not dt and val:
                try: dt = du_parse(str(val), fuzzy=True)
                except: pass
            if dt and (earliest is None or dt < earliest): earliest = dt
            
    return keys, earliest

# --- CORE LOGIC: Expedia PDF Parser (HUNTER MODE V2) ---
# Replaces the old parse_expedia_pdf_to_ws but outputs to an Excel sheet structure
def parse_expedia_hunter_to_ws(pdf_stream, out_wb, sheet_name):
    ws = out_wb.create_sheet(sheet_name)
    headers = ["Booking Type","Reservation ID","Check-In Date","Nights","Guest Name","Currency","Amount Before Tax","Tax","Total Amount Due"]
    ws.append(headers)
    
    keys = set()
    earliest = None
    
    with pdfplumber.open(pdf_stream) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"
            
    # Cleaning
    cleaned = full_text.replace('\n', ' ').replace('"', ' ').replace(',', '')
    cleaned = re.sub(r'\s+', ' ', cleaned)
    
    # Hunter Regex
    matches = re.finditer(r'(Expedia Collect|Hotel Collect)\s.*?(\d{8,15})\s.*?(\d{1,2}-[A-Za-z]{3}-\d{4})', cleaned, re.IGNORECASE)
    
    for m in matches:
        bt = m.group(1)
        rid = m.group(2)
        dt_txt = m.group(3)
        
        # Determine Amounts
        end_pos = m.end()
        window = cleaned[end_pos:end_pos+150]
        prices = re.findall(r'(\d+\.\d{2})', window)
        
        amt, tax, tot = "0.00", "0.00", "0.00"
        if prices:
            tot = prices[-1] # Assume last is total
            if len(prices) >= 2: amt = prices[-2]
            else: amt = tot
            
        try: dt = du_parse(dt_txt)
        except: dt = None
        
        if dt:
            if earliest is None or dt < earliest: earliest = dt
            
        keys.add(norm_key(rid))
        
        # Guest Name (Placeholder)
        guest = "Guest"
        
        ws.append([bt, rid, dt, "1", guest, "THB", amt, tax, tot])
    
    # Format Date Col
    for r in range(2, ws.max_row+1):
        ws.cell(row=r, column=3).number_format = "yyyy-mm-dd"
        
    return ws, 2, 2, earliest, keys

# --- MAIN RECONCILIATION PROCESS ---

def process_reconciliation(hot_file, ota_file):
    # Setup Output Workbook
    out_wb = Workbook()
    out_wb.remove(out_wb.active) # Remove default sheet
    
    # 1. Load Hoteliers
    hot_wb = load_workbook(hot_file, data_only=True)
    hot_ws_src = hot_wb.active
    hot_info = collect_hoteliers(hot_ws_src) # Scan for periods/keys
    
    # 2. Load/Parse OTA
    ota_filename = ota_file.name
    ota_name = os.path.splitext(ota_filename)[0]
    is_pdf_file = ota_filename.lower().endswith('.pdf')
    
    if is_pdf_file:
        # Use Hunter Parser
        com_ws, res_col_idx, data_start, com_earliest, com_keys = parse_expedia_hunter_to_ws(
            ota_file, out_wb, trim_sheet(ota_name))
        com_settings = {"header_row": 1, "reservation_col_idx": res_col_idx, "data_start_row": data_start}
    else:
        # Excel OTA
        com_wb = load_workbook(ota_file, data_only=True)
        com_ws_src = com_wb.active
        com_ws = out_wb.create_sheet(trim_sheet(ota_name))
        for r in com_ws_src.iter_rows(values_only=True):
            com_ws.append(r)
            
        com_settings = detect_commission_settings(com_ws)
        com_keys, com_earliest = collect_commission_excel(com_ws, com_settings)

    # 3. Determine Period & Target OTA
    hot_filename = hot_file.name
    target_period = choose_period(hot_filename, ota_filename, hot_info["earliest_dep"], com_earliest)
    target_ota_type = infer_ota(ota_filename)
    
    # 4. Build "Hoteliers" Sheet (Filtered by Period & OTA)
    hot_sheet_name = trim_sheet(os.path.splitext(hot_filename)[0])
    hot_ws_out = out_wb.create_sheet(hot_sheet_name, 0) # Put first
    
    # Copy Headers
    for r in range(1, 3):
        hot_ws_out.append([c.value for c in hot_ws_src[r]])
        
    # Filter Rows
    hot_rows_period = hot_info["rows_by_period"].get(target_period, [])
    included_rows = []
    ota_col = hot_info["ota_col"]
    
    if ota_col:
        for r_idx in hot_rows_period:
            val = hot_ws_src.cell(row=r_idx, column=ota_col).value
            # Match OTA Logic (v1.15)
            # If target is Expedia, we match "Expedia", "Hotels.com"
            # If target is Booking, match "Booking.com"
            if not target_ota_type or target_ota_type == "OTA": 
                is_match = True
            else:
                cell_canon = canon_ota(str(val))
                target_canon = canon_ota(target_ota_type)
                is_match = (target_canon == cell_canon) or (target_canon in cell_canon)
            
            if is_match:
                included_rows.append(r_idx)
                
        # Fallback: if filtered result is empty, use all rows for period (Warning case)
        if not included_rows and hot_rows_period:
            included_rows = hot_rows_period
    else:
        included_rows = hot_rows_period

    # Copy Included Rows
    hot_out_keys = set()
    hot_res_col = hot_info["res_col"]
    
    for r_idx in included_rows:
        row_vals = [c.value for c in hot_ws_src[r_idx]]
        hot_ws_out.append(row_vals)
        # Record key
        k = norm_key(hot_ws_src.cell(row=r_idx, column=hot_res_col).value)
        if k: hot_out_keys.add(k)
        
    # Format Dates in Output
    for c in [hot_info["arr_col"], hot_info["dep_col"]]:
        if c: format_dates(hot_ws_out, [c], 3, hot_ws_out.max_row)

    # 5. Calculate Differences
    only_hot = hot_out_keys - com_keys
    only_com = com_keys - hot_out_keys
    
    # --- Booking.com Special Logic: Check Amount == 0 ---
    if "booking" in ota_filename.lower() and not is_pdf_file:
        # Locate Final & Commission columns in OTA sheet
        hdr_row = com_settings["header_row"]
        final_col = None
        comm_col = None
        for c in range(1, com_ws.max_column+1):
            v = str(com_ws.cell(row=hdr_row, column=c).value or "").lower()
            if "final" in v and "amount" in v: final_col = c
            if "commission" in v and "amount" in v: comm_col = c
            
        real_only_com = set()
        for k in only_com:
            # Find the row for this key to check amount
            # This is slow O(N^2) but safe. Optimization: Build map first if needed.
            # For simplicity in this structure:
            row_to_check = None
            col_idx = com_settings["reservation_col_idx"]
            for r in range(com_settings["data_start_row"], com_ws.max_row+1):
                if norm_key(com_ws.cell(row=r, column=col_idx).value) == k:
                    row_to_check = r
                    break
            
            if row_to_check:
                f_val = 0
                c_val = 0
                try: f_val = float(com_ws.cell(row=row_to_check, column=final_col).value or 0)
                except: pass
                try: c_val = float(com_ws.cell(row=row_to_check, column=comm_col).value or 0)
                except: pass
                
                if f_val != 0 and c_val != 0:
                    real_only_com.add(k)
        
        # Update set
        only_com = real_only_com
        # Recalculate only_hot (actually hoteliers doesn't change based on OTA amounts, but intersection might)
        # Logic: If OTA key is ignored (amount 0), it shouldn't flag Hoteliers as missing?
        # Standard v1.15 logic: If it's in OTA but 0, we treat it as "Exists but ignore". 
        # So we remove it from only_hot check? No, if it's in OTA (even 0), it exists.
        # So Hoteliers shouldn't flag red. 
        # But wait, we used `com_keys` which includes ALL OTA keys.
        # So `only_hot = hot - com_keys` is correct (if key exists in OTA, hot is fine).
        # The logic only applies to `only_com` (Extra in OTA) - we don't want to highlight 0-amount cancellations.

    # 6. Apply Highlights
    # Hoteliers Sheet
    for r in range(3, hot_ws_out.max_row+1):
        k = norm_key(hot_ws_out.cell(row=r, column=hot_res_col).value)
        if k in only_hot: # Missing in OTA
            for c in range(1, hot_ws_out.max_column+1): hot_ws_out.cell(row=r, column=c).fill = RED_FILL
        elif k in only_com: # (Shouldn't happen in Hoteliers sheet logic usually, but consistent with v1.15)
            pass 

    # OTA Sheet
    c_col = com_settings["reservation_col_idx"]
    start = com_settings["data_start_row"]
    for r in range(start, com_ws.max_row+1):
        k = norm_key(com_ws.cell(row=r, column=c_col).value)
        if k in only_com: # Missing in Hoteliers (and passed amount check)
             for c in range(1, com_ws.max_column+1): com_ws.cell(row=r, column=c).fill = RED_FILL

    # 7. Create NotMatched Sheet
    nm = out_wb.create_sheet("NotMatched")
    nm["A1"].value = f"Hoteliers only ({len(only_hot)})"; nm["A1"].font = Font(bold=True); nm["A1"].fill = GREEN_FILL
    nm["C1"].value = f"Commission only ({len(only_com)})"; nm["C1"].font = Font(bold=True); nm["C1"].fill = RED_FILL
    nm["A2"].value = "Reservation Number"; nm["A2"].font = Font(bold=True)
    nm["C2"].value = "Reservation Number"; nm["C2"].font = Font(bold=True)
    
    r = 3
    for k in sorted(only_hot):
        nm.cell(row=r, column=1, value=str(k)).fill = GREEN_FILL
        r += 1
    r = 3
    for k in sorted(only_com):
        nm.cell(row=r, column=3, value=str(k)).fill = RED_FILL
        r += 1

    # 8. Create Tools Sheet
    tools = out_wb.create_sheet("Commission_Tools")
    tools.append(["Hoteliers File", hot_filename])
    tools.append(["OTA File", ota_filename])
    tools.append(["Period Detected", target_period])
    tools.append(["OTA Detected", target_ota_type])
    tools.append(["Note", "Generated by Web Reconcile v1.15 (Hunter Mode)"])
    
    return out_wb, target_period, len(only_hot), len(only_com)

# --- WEB UI ---

st.title("🏨 Hotel Reconcile Pro (Web Version)")
st.markdown("""
**โปรแกรมกระทบยอด (Reconcile) ฉบับสมบูรณ์**
* ย้ายจากโปรแกรม exe มาเป็นเว็บ
* รองรับไฟล์ Expedia ที่รูปแบบเพี้ยน (Hunter Mode)
* คงฟีเจอร์การกรอง Period, Channel และการเช็คยอด Booking.com ของเดิมไว้ทั้งหมด
""")

col1, col2 = st.columns(2)
with col1:
    hot_files = st.file_uploader("1. Upload Hoteliers Excel", type=["xlsx"], accept_multiple_files=True)
with col2:
    ota_files = st.file_uploader("2. Upload OTA Files (Excel/PDF)", type=["xlsx", "pdf"], accept_multiple_files=True)

if st.button("🚀 Process Reconcile", type="primary"):
    if not hot_files or not ota_files:
        st.error("Please upload both Hoteliers and OTA files.")
    else:
        # Load all Hoteliers into memory map (Filename -> Bytes) to allow reuse
        # In this logic, we assume 1 OTA file matches 1 Hoteliers file (or best guess)
        # v1.15 Logic: Iterate OTA files, find best matching Hoteliers file
        
        st.info("Processing...")
        
        # Prepare Hoteliers list
        # We need to keep them readable multiple times, so read to bytes
        hot_buffers = []
        for hf in hot_files:
            hot_buffers.append({"name": hf.name, "bytes": hf.getvalue()})
            
        results_zip = []
        
        for of in ota_files:
            ota_name = of.name
            
            # Find best match Hoteliers
            # 1. Same Hotel Code?
            # 2. Matching Period in name?
            ota_hotel = infer_hotel(ota_name)
            ota_period = extract_period_from_name(ota_name)
            
            best_hot = None
            
            # Simple matching strategy: Match Hotel Code first
            candidates = [h for h in hot_buffers if infer_hotel(h["name"]) == ota_hotel]
            if not candidates: candidates = hot_buffers # Fallback to all
            
            # If period found in OTA name, try to find in Hoteliers name
            if ota_period:
                period_matches = [h for h in candidates if ota_period.lower() in h["name"].lower()]
                if period_matches: best_hot = period_matches[0]
            
            if not best_hot and candidates: best_hot = candidates[0]
            
            if best_hot:
                # Wrap in BytesIO for processing
                hot_stream = BytesIO(best_hot["bytes"])
                hot_stream.name = best_hot["name"]
                
                ota_stream = of # Streamlit uploader is already a stream-like
                
                try:
                    wb_result, period, n_hot, n_ota = process_reconciliation(hot_stream, ota_stream)
                    
                    # Save to buffer
                    out_io = BytesIO()
                    wb_result.save(out_io)
                    out_io.seek(0)
                    
                    out_name = f"Reconcile-{ota_hotel}-{ota_name[:10]}-{period}.xlsx"
                    
                    st.success(f"✅ {out_name} | Miss Hot: {n_hot} | Miss OTA: {n_ota}")
                    st.download_button(label=f"📥 Download {out_name}", data=out_io, file_name=out_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    
                except Exception as e:
                    st.error(f"Error processing {ota_name}: {e}")
            else:
                st.warning(f"Could not find matching Hoteliers file for {ota_name}")
